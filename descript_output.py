from fastapi import FastAPI
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
import pandas as pd
import cv2
import os
import math
import random

app = FastAPI()

video_path = 'video/조_우0427_14'
video_name = 'video/조_우0427_14.mp4'

video_dir = os.path.join(os.path.dirname(__file__), "video")
image_dir = os.path.join(os.path.dirname(__file__), video_path)

app.mount("/video", StaticFiles(directory=video_dir), name="video")
app.mount(f"/{video_path}", StaticFiles(directory=image_dir), name="baby_test_video")

video = cv2.VideoCapture(video_name)

fps = math.trunc(video.get(cv2.CAP_PROP_FPS))
request_sec = 2
get_frame = math.trunc(fps * request_sec)

@app.get("/", response_class=HTMLResponse)
async def root():
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Descript Output</title>
        <style>
            html, body {
                height: 100%;
                margin: 0;
                font-family: Arial, sans-serif;
            }
            body {
                display: flex;
                justify-content: center;
                align-items: center;
                background-color: #f9f9f9;
            }
            #container {
                display: flex;
                flex-direction: column;
                align-items: center;
                background: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            }
            #buttons {
                display: flex;
                gap: 15px;
                margin-bottom: 20px;
            }
            button {
                padding: 10px 20px;
                font-size: 16px;
                cursor: pointer;
                border: none;
                border-radius: 5px;
                background-color: #007BFF;
                color: white;
                transition: background-color 0.3s;
            }
            button:hover {
                background-color: #0056b3;
            }
            video {
                border-radius: 8px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.2);
            }
            #progressContainer {
                position: relative;
                width: 640px;
                height: 10px;
                background-color: #ddd;
                border-radius: 5px;
                margin-top: 10px;
                margin-bottom: 10px;
            }
            #progressBar {
                height: 100%;
                width: 0%;
                background-color: #007BFF;
                transition: width 0.2s;
            }
            #timeDisplay {
                width: 640px;
                margin-top: 5px;
                font-size: 14px;
                text-align: right;
                color: #333;
                font-family: monospace;
            }
            #subtitle {
                width: 640px;
                text-align: center;
                font-size: 18px;
                font-weight: bold;
                color: black;
                text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.7);
                margin-top: 8px;
            }
            #frameCountDisplay {
                margin-top: 10px;
                font-size: 16px;
                font-weight: bold;
                color: #555;
            }
            #panorama {
                margin-top: 15px;
                display: flex;
                gap: 10px;
                overflow-x: auto;
                width: 660px;
                padding: 10px;
                border: 1px solid #ccc;
                border-radius: 8px;
                background-color: #f1f1f1;
            }
            #panorama img {
                height: 90px;
                cursor: pointer;
                border-radius: 4px;
                transition: transform 0.2s;
            }
            #panorama img:hover {
                transform: scale(1.1);
                box-shadow: 0 0 6px rgba(0, 0, 0, 0.3);
            }
            .marker {
                position: absolute;
                top: 0;
                bottom: 0;
                height: 20px;
                width: 2px;
                background-color: black;
                opacity: 0.8;
                z-index: 2;
                cursor: pointer;
                overflow: visible;
            }

        </style>
    </head>
    <body>
        <div id="container">
            <h1 style="text-align: center; width: 100%; margin-top: 0; margin-bottom: 20px;">
                아기 행동 분석 표
            </h1>
            <div id="excelTableContainer" style="margin: 40px; width: 660px; height: 660px; overflow: auto;"></div>
            
        </div>

        <div id="container">
            <h1 style="text-align: center; width: 100%; margin-top: 0; margin-bottom: 20px;">
                LG U+ 홈Agent
            </h1>
            <div id="buttons">
                <button onclick="playVideo()">ㅤ▶ 원본 재생ㅤ</button>
                <button onclick="playStepByStep()">ㅤ⏭ 프레임 재생ㅤ</button>
                <button onclick="playFilteredFrames()">ㅤ🎬 장면 전환 영상 재생ㅤ</button>
                <button onclick="stopVideo()">ㅤ❚❚ 영상 일시정지ㅤ</button>
            </div>
            <video id="videoPlayer" width="640" controls>
                <source src="videoname" type="video/mp4">
                Your browser does not support the video tag.
            </video>
            <div id="progressContainer">
                <div id="progressBar"></div>
            </div>
            <div id="timeDisplay">00:00 / 00:00</div>
            <div id="subtitle"></div>
            <div id="frameCountDisplay"></div>
            <div id="panorama"></div>
        </div>


        <script>
            const video = document.getElementById('videoPlayer');
            const progressBar = document.getElementById('progressBar');
            const timeDisplay = document.getElementById('timeDisplay');
            const subtitle = document.getElementById('subtitle');
            const frameCountDisplay = document.getElementById('frameCountDisplay');

            let globalCurrentTime = 0;
            let currentMode = null;
            let currentTask = null;
            let filteredFrames = [];
            let subtitles = [];

            async function stopVideo() {
                currentMode = null;
                video.pause()
                await stopCurrentTask();
            }

            async function stopCurrentTask() {
                if (currentTask && typeof currentTask.cancel === "function") {
                    currentTask.cancel();
                }
                await new Promise(resolve => setTimeout(resolve, 100));
                currentMode = null;
            }
        
            function formatTime(seconds) {
                const min = Math.floor(seconds / 60);
                const sec = Math.floor(seconds % 60);
                return `${String(min).padStart(2, '0')}:${String(sec).padStart(2, '0')}`;
            }
        
            async function playVideo() {
                if (currentMode === "normal") return;
                await stopCurrentTask();
                currentMode = "normal";
                video.currentTime = globalCurrentTime;
                video.play();
            }

        
            async function playStepByStep() {
                if (currentMode === "step") return;
                await stopCurrentTask();
                currentMode = "step";
            
                let cancel = false;
                currentTask = { cancel: () => cancel = true };
            
                video.pause();
                video.currentTime = globalCurrentTime;
            
                while (video.currentTime < video.duration) {
                    if (cancel) return;
                    await new Promise(resolve => setTimeout(resolve, 1000));
                    if (cancel) return;
                    video.currentTime += request_sec;
                    globalCurrentTime = video.currentTime;
                    const subtitleText = getSubtitleAtTime(video.currentTime);
                    subtitle.textContent = subtitleText || '';
                    await new Promise(resolve => video.onseeked = () => resolve());
                }
            }
        
            async function loadFilteredFrames() {
                const response = await fetch('/frames_with_o');
                const data = await response.json();
                filteredFrames = data.frames;
                frameCountDisplay.textContent = `장면 전환 수: ${data.frame_count}`;
        
                const panorama = document.getElementById('panorama');
                panorama.innerHTML = '';
        
                data.image_urls.forEach((url, index) => {
                    const clickedTime = filteredFrames[index];
                    const subtitleText = getSubtitleAtTime(clickedTime);
                
                    // 이미지와 시간 표시를 감싸는 컨테이너 생성
                    const container = document.createElement('div');
                    container.style.display = 'inline-block';
                    container.style.textAlign = 'center';
                    container.style.margin = '5px';
                
                    // 이미지 생성
                    const img = document.createElement('img');
                    img.src = url;
                    img.style.display = 'block';
                    img.style.cursor = 'pointer';
                    img.addEventListener('click', async () => {
                        const clickedTime = filteredFrames[index];
                        
                        if (currentMode === "filtered") {
                            if (currentTask && typeof currentTask.cancel === "function") {
                                currentTask.cancel();
                            }
                            video.currentTime = clickedTime;
                            let cancel = false;
                            currentTask = { cancel: () => cancel = true };
                            for (let i = index; i < filteredFrames.length; i++) {
                                if (cancel) return;
                                video.currentTime = filteredFrames[i];
                                await new Promise(resolve => video.onseeked = () => resolve());
                                const subText = getSubtitleAtTime(video.currentTime);
                                subtitle.textContent = `${subText} (${i + 1}/${filteredFrames.length})`;
                                await new Promise(resolve => setTimeout(resolve, 1000));
                            }
                            currentTask = null;
                        } else if (currentMode === "step") {
                            if (currentTask && typeof currentTask.cancel === "function") {
                                currentTask.cancel();
                            }
                            video.pause();
                            video.currentTime = clickedTime;
                    
                            let cancel = false;
                            currentTask = { cancel: () => cancel = true };
                    
                            const interval = 2;
                    
                            while (video.currentTime < video.duration) {
                                if (cancel) return;
                                await new Promise(resolve => setTimeout(resolve, 1000));
                                if (cancel) return;
                                video.currentTime += interval;
                
                                const subtitleText = getSubtitleAtTime(video.currentTime);
                                subtitle.textContent = subtitleText || '';
                                await new Promise(resolve => video.onseeked = () => resolve());
                            }
                            currentTask = null;
                        } else if (currentMode === "normal") {
                            if (currentTask && typeof currentTask.cancel === "function") {
                                currentTask.cancel();
                            }
                            video.play();
                            video.currentTime = clickedTime;
                    
                            let cancel = false;
                            currentTask = { cancel: () => cancel = true };
                        } else {
                            await stopCurrentTask();
                            currentMode = null;  
                            currentTask = null;
                            
                            video.pause();      
                            video.currentTime = clickedTime;    
                
                            const subText = getSubtitleAtTime(video.currentTime);
                            subtitle.textContent = `${subText} (${index+1}/${filteredFrames.length})`;
                        }
                    });
                
                    // 시간 텍스트 생성
                    const timeLabel = document.createElement('span');
                    let time_text = formatTime(clickedTime);
                    let time_descript = subtitleText;
                    let res = time_text + "\\n" + String(time_descript);
                    timeLabel.textContent = res;
                    timeLabel.style.fontSize = '0.8em';
                    timeLabel.style.color = '#555';
                
                    // 컨테이너에 이미지와 텍스트 추가
                    container.appendChild(img);
                    container.appendChild(timeLabel);
                    panorama.appendChild(container);
                });

                addMarkersToProgressBar(filteredFrames);
            }
        
            async function playFilteredFrames() {
                if (currentMode === "filtered") return;
                await stopCurrentTask();
                currentMode = "filtered";

                let cancel = false;
                currentTask = { cancel: () => cancel = true };

                if (filteredFrames.length === 0) {
                    await loadFilteredFrames();
                }

                video.pause();
                video.currentTime = globalCurrentTime;

                for (let i = 0; i < filteredFrames.length; i++) {
                    if (cancel) return;

                    if (globalCurrentTime <= filteredFrames[i]) {
                        globalCurrentTime = filteredFrames[i];
                        video.currentTime = globalCurrentTime;
                        await new Promise(resolve => video.onseeked = () => resolve());
            
                        const subText = getSubtitleAtTime(video.currentTime);
                        subtitle.textContent = `${subText} (${i + 1}/${filteredFrames.length})`;

                        await new Promise(resolve => setTimeout(resolve, 1000));
                    }
                }

                currentMode = null;
                currentTask = null;
            }

        
            async function loadSubtitles() {
                const response = await fetch('/subtitles');
                const data = await response.json();
                subtitles = data.subtitles;
            }
        
            function getSubtitleAtTime(currentTime) {
                for (let i = 0; i < subtitles.length; i++) {
                    const sub = subtitles[i];
                    if (currentTime >= sub.start && currentTime <= sub.end) {
                        return sub.text;
                    }
                }
                return "";
            }
        
            video.addEventListener('timeupdate', () => {
                globalCurrentTime = video.currentTime;
                const percent = (video.currentTime / video.duration) * 100;
                progressBar.style.width = percent + "%";

                if (!isNaN(video.duration)) {
                    const current = formatTime(video.currentTime);
                    const total = formatTime(video.duration);
                    timeDisplay.textContent = `${current} / ${total}`;
                }

                // 현재 프레임 계산 (예: 30fps 기준)
                const fps = 30;
                const currentFrame = Math.floor(video.currentTime * fps);
                

            });

            function addMarkersToProgressBar(frames) {
                document.querySelectorAll('.marker').forEach(m => m.remove());
                if (!video.duration || video.duration === Infinity) return;

                const tooltip = document.createElement('div');
                tooltip.id = 'markerTooltip';
                tooltip.style.position = 'absolute';
                tooltip.style.padding = '4px 8px';
                tooltip.style.fontSize = '12px';
                tooltip.style.backgroundColor = 'black';
                tooltip.style.color = 'white';
                tooltip.style.borderRadius = '4px';
                tooltip.style.pointerEvents = 'none';
                tooltip.style.opacity = 0;
                tooltip.style.transition = 'opacity 0.2s';
                tooltip.style.zIndex = 5;
                document.body.appendChild(tooltip);

                frames.forEach(time => {
                    const marker = document.createElement('div');
                    marker.classList.add('marker');
                    const percent = (time / video.duration) * 100;
                    marker.style.left = `${percent}%`;

                    marker.addEventListener('click', async (e) => {
                        await stopCurrentTask();
                        currentMode = null;
                        video.pause();
                        video.currentTime = time;
                        subtitle.textContent = ``;
                    });

                    marker.addEventListener('mouseover', (e) => {
                        let currentTime = formatTime(time);
                        const subText = getSubtitleAtTime(time) + ' 현재 시간: ' + currentTime ;
                        tooltip.textContent = subText;
                        const rect = e.target.getBoundingClientRect();
                        tooltip.style.left = `${rect.left + window.scrollX}px`;
                        tooltip.style.top = `${rect.top + window.scrollY - 28}px`;
                        tooltip.style.opacity = 1;
                    });

                    marker.addEventListener('mouseout', () => {
                        tooltip.style.opacity = 0;
                    });

                    progressContainer.appendChild(marker);
                });
            }
        
            loadFilteredFrames();
            loadSubtitles();
            
            async function loadExcelData() {
                const response = await fetch('/excel_data');
                const result = await response.json();

                const container = document.getElementById('excelTableContainer');
                container.innerHTML = '';

                const table = document.createElement('table');
                table.style.borderCollapse = 'collapse';
                table.style.width = '100%';
                table.style.fontSize = '14px';
                table.style.border = '1px solid #ccc';

                const thead = document.createElement('thead');
                const headerRow = document.createElement('tr');
                result.columns.forEach(col => {
                    const th = document.createElement('th');
                    th.textContent = col;
                    th.style.padding = '8px';
                    th.style.border = '1px solid #ccc';
                    th.style.backgroundColor = '#f0f0f0';
                    th.style.textAlign = 'left';
                    headerRow.appendChild(th);
                });
                thead.appendChild(headerRow);
                table.appendChild(thead);

                const tbody = document.createElement('tbody');
                result.data.forEach(row => {
                    const tr = document.createElement('tr');
                    row.forEach((cell, colIndex) => {
                        const td = document.createElement('td');
                        td.style.padding = '8px';
                        td.style.border = '1px solid #ccc';

                        if (colIndex === 0 && typeof cell === 'string' && cell.trim() !== '') {
                            const img = document.createElement('img');
                            img.src = cell.trim();
                            img.style.height = '80px';
                            img.style.objectFit = 'contain';
                            td.appendChild(img);
                        } else {
                            td.textContent = cell !== null ? cell : '';
                        }

                        tr.appendChild(td);
                    });
                    tbody.appendChild(tr);
                });
                table.appendChild(tbody);

                container.appendChild(table);
            }

            loadExcelData();

        </script>
    </body>
    </html>
    """.replace("videoname",video_name).replace("request_sec", str(request_sec))

request_sec = 2

@app.get("/frames_with_o")
async def get_frames_with_o():
    
    excel_path = "./excel/description.xlsx"
    wb = load_workbook(excel_path)
    ws = wb.active
    frames_o = []
    image_urls = []
    frame_count = 0
    for row in ws.iter_rows(min_row=2):
        mark = row[4].value
        if mark == "O":
            time_str = row[1].value
            if time_str:
                h, m, s = map(int, time_str.split(":"))
                file_time = f"{h:02d}_{m:02d}_{s:02d}"
                total_seconds = h * 3600 + m * 60 + s
                frames_o.append(total_seconds)
                image_filename = f"baby_test_{file_time}.jpg"
                image_urls.append(f"/{video_path}/{image_filename}?v={random.randint(1, 1_000_000)}")
                frame_count += 1
    return JSONResponse(content={"frames": frames_o, "frame_count": frame_count, "image_urls": image_urls})

@app.get("/subtitles")
async def get_subtitles():
    excel_path = "./excel/description.xlsx"
    wb = load_workbook(excel_path)
    ws = wb.active

    subtitles = []
    for row in ws.iter_rows(min_row=2):
        time_str = row[1].value
        text = row[2].value
        if time_str and text:
            h, m, s = map(int, time_str.split(":"))
            start_time = h * 3600 + m * 60 + s
            end_time = start_time + 1
            subtitles.append({"start": start_time, "end": end_time, "text": text})

    return JSONResponse(content={"subtitles": subtitles})

@app.get("/excel_data")
async def get_excel_data():
    excel_path = "./excel/description.xlsx"
    result_sheet = pd.read_excel(excel_path, sheet_name='result', usecols= [1,2,5])

    result_sheet = result_sheet.replace([float('inf'), float('-inf')], pd.NA)
    result_sheet = result_sheet.fillna('')

    result_wb = load_workbook(excel_path)
    result_ws = result_wb.active
    image_urls = []
    for row in result_ws.iter_rows(min_row=2):
        mark = row[4].value
        time_str = row[1].value
        if mark == "O":
            h, m, s = map(int, time_str.split(":"))
            file_time = f"{h:02d}_{m:02d}_{s:02d}"
            image_filename = f"baby_test_{file_time}.jpg"
            image_urls.append(f"/{video_path}/{image_filename}")
    
    result_sheet.insert(0, "이미지", image_urls)

    return JSONResponse(content={
        "columns": list(result_sheet.columns),
        "data": result_sheet.values.tolist()
    })