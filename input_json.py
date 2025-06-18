from openpyxl import load_workbook
import json

root_path = "."
file_name = "LG U+ 홈Agent PoC_GT"
sheet_name = "류_상250426_11"
wb = load_workbook(f"{root_path}/gt_xlsx/{file_name}.xlsx")
ws = wb[sheet_name]

# 헤더 추출 (1행)
headers = [cell.value for cell in ws[1]]

# 데이터 추출 (2행부터)
json_data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    row_dict = dict(zip(headers, row))
    row_dict["TIMESTAMP"] = row_dict["TIMESTAMP"].strftime("%H:%M:%S")
    if row_dict["ACTIONTYPE"] == "NONE":
        row_dict["IS_EXIST"] = "X"
    else:
        row_dict["IS_EXIST"] = "O"

    json_data.append(row_dict)


# 결과 출력
# for data in json_data:
#     print(data)
#     exit()

# 결과 저장
dst_json_file = f"{root_path}/gt/{sheet_name}.json"
with open(dst_json_file, 'w', encoding='utf-8') as f:
    json_data = json.dumps(json_data, ensure_ascii=False, indent=4)
    f.write(json_data)

#액션 Type 종류 : SLEEP, AWAKE, MOVE, Cry, NONE
#descript : 아기가 자는 중입니다, 아기가 깨어 있습니다, 아기가 움직이는 중입니다, 아기가 울고 있습니다, 아기가 침대에 존재합니다/ 아기가 침대에 존재하지 않습니다.
