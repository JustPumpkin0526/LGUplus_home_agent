from PIL import Image

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from sentence_transformers import SentenceTransformer

from fastapi import FastAPI
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

from inference_client import infer_from_server_with_image_object
from textsimilarity import get_similarity

import json
import time
import cv2
import os
import math

code_start = time.time()

#엑셀의 셀을 이미지 크기에 맞게 변환하기 위해 크기를 조정하는 코드
def get_img_size(img_width, img_height):
    col_width = img_width*31.5/252.19
    row_height = img_height*149.1/198.96
    return (col_width, row_height)

#폴더 내부 요소 전부 삭제
def DeleteAllFiles(folderpath):
    if os.path.exists(folderpath):
        for file in os.scandir(folderpath):
            os.remove(file.path)
        return 'Remove All File'
    else:
        return 'Directory Not Found'

#엑셀 연동하기
write_wb = Workbook()
write_ws = write_wb.active

#영상 불러오기
#김_현250428_9
#류_상250426_11
#류_상250427_14
#조_우0427_14
#baby_test_video(LGU+1)
#baby_test_video(LGU+2)
#baby_test_video(LGU+3)

filepath = 'video/류_상250426_11.mp4'
video = cv2.VideoCapture(filepath)

if not video.isOpened():
    print("열리지 않는 파일")
    exit(0)

#프레임별 추출한 이미지 담을 폴더 생성
try:
    if not os.path.exists(filepath[:-4]) and not os.path.exists(filepath[:-4]+"_result"):
        os.makedirs(filepath[:-4])
    else:
        DeleteAllFiles(filepath[:-4])
except OSError:
    print('Error: Already created directory. ' + filepath[:-4])

#유사도 측정 모델 불러오기
device = "cuda:0"
model_sentence = SentenceTransformer('/model_clone/ko-sroberta-multitask',device=device)

#GT 불러오기

json_file = "./gt/류_상250426_11.json"
with open(json_file, 'r', encoding='utf-8') as f:
    GT_json = json.load(f)

count = 1
GT_Sleep_O_count = 0
GT_Sleep_X_count = 0
GT_Awake_O_count = 0
GT_Awake_X_count = 0
GT_SLEEP_O = 0
GT_SLEEP_X = 0
fps = math.trunc(video.get(cv2.CAP_PROP_FPS))
request_sec = 5
get_frame = math.trunc(fps * request_sec)

write_ws.append(["IMAGE","TIMESTAMP","DESCRIPT","SIMILARITY_SCORE","SLEEP_SCORE","SLEEP_CHANGE", "SLEEP_DESCRIPT", "SLEEP_PREDICTION", "SLEEP_GT", "SLEEP_GT_RESULT" , "TOTAL_SCORE", "SLEEP_CHANGE_GT"])
#프레임별 이미지 추출 및 타임스탬프, descript 받아오기
folder_name = filepath[:-4]

excel_folder_path = f"excel/{folder_name[6:]}.xlsx"

GT_change_count = 0

#GT 장면 전환 수 체크
for GT_DICT in GT_json:
    if GT_DICT["CHANGE"] == 'O':
        GT_change_count += 1

while(video.isOpened()):
    ret, image = video.read()
    frame = int(video.get(1))
    if frame == 1:
        frame = 0    
    if not ret:
        break

    if(frame % get_frame == 0):
        start = time.time()
        
        #현재 프레임에 fps를 나눠 현재 몇초인지를 알 수 있음
        frame_second = int(frame) // fps
        second = frame_second%60
        minute = (frame_second//60)%60 
        hour = frame_second//60//60
 
        #타임 스탬프 구하기
        Time = f"{hour:02d}:{minute:02d}:{second:02d}"
        file_time = f"{hour:02d}_{minute:02d}_{second:02d}"

        file_name = f"baby_test_{file_time}.jpg"
        cv2.imwrite(folder_name + "/%s" % file_name, image)

        count += 1  

        image_path = folder_name + "/%s" %file_name
        request_img = Image.open(image_path)
        # 그라디오 사이트 URL : https://62283cc7c6508eb26d.gradio.live/

        # 고려 중인 프롬프트 : 아기가 잠을 자는 중인가요? 잠을 자는 중이 아니라면 무슨 행동을 하고있나요?, 아기 상황을 설명해주세요, 아기가 자는 중인가요? ‘네’ 혹은 ‘아니요’로 대답을 통일하세요, 아기의 상태와 행동을 설명헤주세요, 
        # 아기가 잠을 자고 있나요? 자고 있지 않다면 무슨 행동을 하는 중인가요?, 아기가 잠을 자고 있나요? 아기가 깨어 있다면 무슨 행동을 하는 중인가요?, 아기의 수면 여부를 설명하시오., 아기는 지금 자고 있습니까?

        #시도해본 프롬프트 : 아기가 자는 중인가요? 아니면 깨어 있나요?, 아기의 수면 상태는 어떤가요?, 아기가 깨어있나요? 아니면 자는 중인가요?, 아기의 상태는 어떤가요?, 아기의 수면 여부는 어떤가요?, 아기는 무슨 상태인가요?,아기가 자는 중인가요?, 아기는 자고 있습니까?,
        #아기가 잠을 자고 있나요? 아기가 깨어 있다면 무슨 행동을 하는 중인가요?, 아기가 자고 있나요? 아니면 아기가 깨어 있나요?
        start_time = time.time()
        result = infer_from_server_with_image_object("http://172.16.8.52:8000", request_img, "아기가 자고 있나요? 아니면 아기가 깨어 있나요?","Llama3.2-VIX-M-3B-KO")
        end_time = time.time()
        
        img = XLImage(image_path)
        img.width = 300
        img.height = 200

        image_cell = 'A'+ str(count)
        descript = result.get('result')
    
        write_ws.add_image(img, image_cell)
        write_ws.append({'B':Time, 'C':descript})

        width, height = get_img_size(img.width, img.height)

        #수면 여부 체크 시작
        keywards = ["자고", "자는", "수면", "잠"]
        for keyword in keywards:
            if keyword in descript:
                check = True
                break
            else:
                check = False        

        if check == True:
            sleep_score = max(get_similarity(model_sentence, descript, "아기가 자고 있습니다."), get_similarity(model_sentence, descript, "아기가 자는 중입니다."), get_similarity(model_sentence, descript, "아기가 수면 중입니다."), get_similarity(model_sentence, descript, "아기가 잠에 들었습니다."))
            write_ws['E'+str(count)] = sleep_score
            if sleep_score >= 0.7:
                sleep_descript = "아기가 자는 중입니다."
                write_ws['H'+str(count)] = "O"
            else :
                sleep_descript = "아기가 깨어 있습니다."
                write_ws['H'+str(count)] = "X"
        else :
            sleep_descript = "아기가 깨어 있습니다."
            write_ws['H'+str(count)] = "X"
        
        write_ws['G'+str(count)] = sleep_descript
        #수면 여부 체크 끝

        #GT 체크
        for GT_DICT in GT_json:
            if GT_DICT["TIMESTAMP"] == Time:
                write_ws['I'+str(count)] = GT_DICT["IS_SLEEP"]
                if GT_DICT["IS_SLEEP"] == "O":
                    GT_SLEEP_O += 1
                    if sleep_descript == "아기가 자는 중입니다.":
                        write_ws['J'+str(count)] = "O"
                        GT_Sleep_O_count += 1
                    else :
                        write_ws['J'+str(count)] = "X"
                        GT_Sleep_X_count += 1
                else :
                    GT_SLEEP_X += 1
                    if sleep_descript == "아기가 깨어 있습니다.":
                        write_ws['J'+str(count)] = "O"
                        GT_Awake_O_count += 1
                    else :
                        write_ws['J'+str(count)] = "X"
                        GT_Awake_X_count += 1
                if frame_second == 0:
                    for i in range(frame_second, frame_second + (request_sec // 2 + 1)):
                        if len(GT_json) <= i:
                            break
                        if GT_json[i]["CHANGE"] == 'O':
                            if (GT_json[i]["IS_SLEEP"] == "O" and sleep_descript == "아기가 자는 중입니다.") or (GT_json[i]["IS_SLEEP"] == "X" and sleep_descript == "아기가 깨어 있습니다."):
                                write_ws['L'+str(count)] = GT_json[i]["CHANGE"]
                    break
                else:
                    for i in range(frame_second - (request_sec // 2 + 1), frame_second + (request_sec // 2 + 1)):
                        if len(GT_json) <= i:
                            break
                        if GT_json[i]["CHANGE"] == 'O':
                            if (GT_json[i]["IS_SLEEP"] == "O" and sleep_descript == "아기가 자는 중입니다.") or (GT_json[i]["IS_SLEEP"] == "X" and sleep_descript == "아기가 깨어 있습니다."):
                                write_ws['L'+str(count)] = GT_json[i]["CHANGE"]
                    break
        #GT 체크 끝

        #소요 시간 및 A열 크기 지정
        write_ws.column_dimensions['A'].width = width
        write_ws.row_dimensions[count].height = height
        end = time.time()
        print(f"vlm 요청 시간은 {end_time - start_time:.5f}초 입니다.\n")
        print(f"{end - start:.5f} sec\n")

#GT 정확도 판단
write_ws['K2'] = f"정확도 : {((GT_Sleep_O_count+GT_Awake_O_count) / count) * 100}"
write_ws['K3'] = f"SLEEP GT의 개수 : {GT_SLEEP_O}, AWAKE GT의 개수 : {GT_SLEEP_X}\nSleep GT 정답 개수 : {GT_Sleep_O_count}, Awake GT 정답 개수 : {GT_Awake_O_count},\nSleep GT 오답 개수 : {GT_Sleep_X_count}, Awake GT 오답 개수 : {GT_Awake_X_count},\n정답 총합 개수 : {GT_Awake_O_count + GT_Sleep_O_count}, 오답 총합 개수 : {GT_Awake_X_count + GT_Sleep_X_count}"


write_wb.save(filename=excel_folder_path)
video.release()

read_wb = load_workbook(filename=excel_folder_path)
read_ws = read_wb["Sheet"]
prediction_change_count = 0
for des_count in range(count, 1, -1):
    if (des_count == 1):
        write_ws['D'+str(des_count)] = 0
        write_ws['F'+str(des_count)] = "O"
        break
    else:
        descript_value = str(read_ws['G'+str(des_count)].value)
        descript_check = str(read_ws['G'+str(des_count-1)].value)
    score = get_similarity(model_sentence, descript_value, descript_check)
    #sleep_score = get_similarity(model_sentence, descript_value, "아기가 자는 중입니다.")
    write_ws['D'+str(des_count)] = score
    #write_ws['G'+str(des_count)] = sleep_score

    if score > 0.65:
        write_ws['F'+str(des_count)] = "X"
    else:
        write_ws['F'+str(des_count)] = "O"
        prediction_change_count += 1
    #if sleep_score > 0.75:
    #    write_ws['F'+str(des_count)] = "아기가 자는 중입니다."
    #else:
    #    write_ws['F'+str(des_count)] = "아기가 깨어 있습니다."

result_ws = write_wb.create_sheet('result')
result_ws.append(["IMAGE","TIMESTAMP","DESCRIPT","SIMILARITY_SCORE","SLEEP_SCORE","SLEEP_CHANGE", "SLEEP_DESCRIPT", "SLEEP_PREDICTION", "SLEEP_GT", "SLEEP_GT_RESULT" , "TOTAL_SCORE", "SLEEP_CHANGE_GT"])
#"아기 수면 여부", "수면 스코어"
cell_num = 1
count_img = -1
result_count = 0
result_GT_count = 0
change_count = 0
file_list = os.listdir(folder_name + "/")
for row in write_ws.iter_rows():
    if row[5].value == 'O':
        result_count += 1
        if row[11].value == 'O':
            change_count += 1
        if row[9].value == 'O' :
            result_GT_count += 1
            
        cell_num +=1
        file_name = file_list[count_img]
        image_path = folder_name + "/%s" %file_name

        result_img = XLImage(image_path)
        result_img.width = 300
        result_img.height = 200

        result_ws.add_image(result_img, 'A'+str(cell_num))
        result_ws.append((cell.value for cell in row))

        width, height = get_img_size(img.width, img.height)
        result_ws.column_dimensions['A'].width = width
        result_ws.row_dimensions[cell_num].height = height
        
    count_img += 1

result_ws['K2'] = f"GT 정확도 : {(result_GT_count / result_count) * 100}"
result_ws['K3'] = f"예측한 장면 전환 수 : {prediction_change_count}, GT 장면 전환 수 : {GT_change_count}, 올바른 장면 전환 예측의 개수 : {change_count} \n예측에 따른 장면 전환 정확도 {(change_count / prediction_change_count) * 100:.3f}"
write_wb.save(filename=excel_folder_path)
code_end = time.time()

print(f"{(code_end - code_start) // 60}분 {(code_end - code_start) % 60}초")